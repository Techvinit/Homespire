import openpyxl

data_file= openpyxl.load_workbook("Book1.xlsx")
raw_data= data_file["Sheet1"]
print(raw_data.max_row)

for product_row in range(2, raw_data.max_row + 1):

    supplier = raw_data.cell(product_row, 3).value
    print(supplier)
 


